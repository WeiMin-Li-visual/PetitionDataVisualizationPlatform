import json
import os
import xlrd
import xlwt
import random
import openpyxl

#按字典value值排序
def dict_sort_by_value(mydict, reverse=True):
    return sorted(mydict.items(), key=lambda kv: (kv[1], kv[0]), reverse=reverse)

#各个类别行数统计
#input：excel目录，excel名称，诉求在表格中哪一列，分类部门值在诉求表格中哪一列
#output：各个类别行数的字典，按频次从高到底排序
def get_lines_count(dir, excel_name, request_Columns, department_Columns):
    lines_count = {}
    data_path = os.path.join(dir, excel_name)
    # 打开文件
    workbook = xlrd.open_workbook(data_path)
    index = workbook.sheet_names()[0]
    sheet = workbook.sheet_by_name(index)
    # 遍历
    nrows = sheet.nrows
    for index in range(1, nrows):
        row_list = sheet.row_values(index)
        # 获取逐条投诉数据
        # request = row_list[request_Columns]
        department = row_list[department_Columns]
        # #机构主题类别标签统计
        if department not in lines_count.keys():
            lines_count[department] = 1
        else:
            lines_count[department] = lines_count[department] + 1

    # 机构主题类别标签统计排序
    lines_count = dict(dict_sort_by_value(lines_count))
    return lines_count

#获取标签对应的数值，按照字典中行数进行排序给定标签对应值，将少于对应阈值的归并为其他类别
#input：统计得到的各类别行数，少数行类别判断阈值，'其他'类别的标签值
#output：真实的标签值，归并少数行到'其他'后的标签值，少数类别的个数
def get_label_value(lines_count, threshold_value=5, other_value=0):
    real_label_value = {}
    label_value = {}
    classNum_low_than_threshold = 0
    index = 0
    for department in lines_count.keys():
        real_label_value[department] = index
        label_value[department] = index
        if lines_count[department] <= threshold_value:
            #统计没超过阈值的类别
            classNum_low_than_threshold = classNum_low_than_threshold + 1
            #归并时，应该将少数类别的标签定为同‘其他’类别一致的标签值
            label_value[department] = other_value
        index = index + 1
    return real_label_value, label_value, classNum_low_than_threshold

# 数据格式转换，excel转换成模型需要的txt格式
#input：excel目录，excel名称，转换格式后txt保存目录，转换格式后txt保存名称，诉求在表格中哪一列，分类部门值在诉求表格中哪一列，统计得到的各个类别的行数，统计得到的各个类别的标签值
#output：转换格式后存储的txt文件，分训练+验证的txt，和test的txt，模型需要test数据集中包含所有标签
def change_data_form(dir, excel_name, save_dir, save_name, request_Columns, department_Columns, label_value):
    lines_count = {}
    data_path = os.path.join(dir, excel_name)
    train_val_txt_save_path = os.path.join(save_dir, save_name + '.txt')
    test_txt_save_path = os.path.join(save_dir, save_name + '_test.txt')
    train_val_Write = open(train_val_txt_save_path, 'r+', encoding='utf-8')
    testWrite = open(test_txt_save_path, 'r+', encoding='utf-8')
    select_label = []
    # 打开文件
    workbook = xlrd.open_workbook(data_path)
    index = workbook.sheet_names()[0]
    sheet = workbook.sheet_by_name(index)
    # 遍历
    nrows = sheet.nrows
    for index in range(1, nrows):
        row_list = sheet.row_values(index)
        # 获取逐条投诉数据
        request = row_list[request_Columns]
        department = row_list[department_Columns]
        label = label_value[department]
        request = request.replace('\n', '')
        request = request.replace('\r', '')
        request = request.replace('\t', '')
        request = request.replace(' ', '')
        if label not in select_label:
            testWrite.write(request + '\t' + str(label) + '\n')
            select_label.append(label)
        else:
            train_val_Write.write(request + '\t' + str(label) + '\n')
    train_val_Write.close()
    testWrite.close()

# 追加转换关键词为训练集
def change_keywords(keyword_dir, keyword_name, save_dir, save_name, label_value):
    txt_keyword_path = os.path.join(keyword_dir, keyword_name + '.txt')
    txt_save_path = os.path.join(save_dir, save_name + '_keywords.txt')
    # 追加寻找到的部门匹配关键词，到训练集行当中
    additionalWrite = open(txt_save_path, 'r+', encoding='utf-8')
    with open(txt_keyword_path, 'r', encoding='utf-8') as file_to_read:
        while True:
            lines = file_to_read.readline()  # 整行读取数据
            if not lines:
                break
            department = lines.split(':')[0]
            wordList = lines.split(':')[1].split(',')
            label = label_value[department]
            for word in wordList:
                txt = word.replace('\n', '')#末尾有个\n换行
                additionalWrite.write(txt + '\t' + str(label) + '\n')
    additionalWrite.close()

# 获取class
def write_class(class_dir, class_name, label_value):
    txt_class_path = os.path.join(class_dir, class_name + '.txt')
    classWrite = open(txt_class_path, 'r+', encoding='utf-8')
    for department in label_value.keys():
        classWrite.write(department + '\n')
    classWrite.close()

# 大表格读写，剔除指定类别='其他'的行
def remove_lines_by_className(dir, excel_name, save_dir, save_name, className='其他'):
    file_path = os.path.join(dir, excel_name)
    save_path = os.path.join(save_dir, save_name)

    # 打开文件
    workbook = openpyxl.load_workbook(file_path)
    sheetnames = workbook.get_sheet_names()
    sheet = workbook.get_sheet_by_name(sheetnames[0])

    ## 或者如果你只是想创建一张空表
    copy_workbook = openpyxl.Workbook()
    # 创建一个sheet
    copy_sheet = copy_workbook.create_sheet(index=0)

    # 遍历
    rows = sheet.max_row
    clos = sheet.max_column
    print('rows:', rows)
    print('clos:', clos)
    index = 1
    for row in range(1, rows + 1):
        label = sheet.cell(row, 6).value
        label = label.replace('\n', '')
        label = label.replace('\r', '')
        label = label.replace('\t', '')
        label = label.replace(' ', '')
        if label.strip() != '其他':
            copy_sheet.cell(index, 1).value = sheet.cell(row, 1).value
            copy_sheet.cell(index, 2).value = sheet.cell(row, 2).value
            copy_sheet.cell(index, 3).value = sheet.cell(row, 3).value
            copy_sheet.cell(index, 4).value = sheet.cell(row, 4).value
            copy_sheet.cell(index, 5).value = sheet.cell(row, 5).value
            copy_sheet.cell(index, 6).value = sheet.cell(row, 6).value
            index += 1
        print(row)

    copy_workbook.save(save_path)
    print('finish...')

# 数据集划分
def split_dataset(dir, txt_name, train_name, val_name, test_name, total_num, train_num, val_num, test_num):
    data_path = os.path.join(dir, txt_name)
    train_path = os.path.join(dir, train_name)
    val_path = os.path.join(dir, val_name)
    test_path = os.path.join(dir, test_name)
    train_Write = open(train_path, 'r+', encoding='utf-8')
    val_Write = open(val_path, 'r+', encoding='utf-8')
    test_Write = open(test_path, 'r+', encoding='utf-8')
    #生成行下标list
    index_list = []
    for index in range(total_num):
        index_list.append(index)
    #打乱行下标list，相当于打乱数据
    random.shuffle(index_list)
    #乱的数据分别填充，进行split
    index = 0
    for line in open(data_path, encoding='utf-8'):
        if index < val_num:
            val_Write.write(line)
        elif index < (val_num + test_num):
            test_Write.write(line)
        else:
            train_Write.write(line)
        index += 1
    train_Write.close()
    val_Write.close()
    test_Write.close()
    print('finished...')

# 测试剔除类别函数remove_lines_by_className
# remove_lines_by_className('../static/data/', 'data_excel1.xlsx', '../static/data/', 'data_excel2.xlsx', className='其他')

# 测试数据集划分函数split_dataset
# split_dataset('../static/data/', 'text_label.txt', 'train.txt', 'val.txt', 'test.txt', 373326, 299120, 37376, 36830)    

# # 测试函数 get_lines_count
line_count_dict = get_lines_count('../static/data/', 'data_excel1.xlsx', 1, 5)
print(line_count_dict)
print(len(line_count_dict))

# 测试 get_label_value函数
real_label_value, label_value, classNum_low_than_threshold = get_label_value(line_count_dict)
print(real_label_value)
print(label_value)
print(classNum_low_than_threshold)

# 测试 change_data_form函数
change_data_form('../static/data/', 'data_excel1.xlsx', '../static/data/', 'text_label', 1, 5, real_label_value)

# 测试 change_keywords函数
change_keywords('../static/data/', 'department_keywords', '../static/data/', 'additional', real_label_value)

# 测试 write_class函数
write_class('../static/data/', 'class', real_label_value)

print('finished.......')
