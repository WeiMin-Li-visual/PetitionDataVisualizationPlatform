import os

import openpyxl
import xlrd
import xlwt
from flask import Flask, render_template, request
import json
import sqlite3
from flask import jsonify
from algorithm import request_extract
from algorithm import emotionalAnalysisOfSingleData
from model.Bert import predict
import random
from flask_socketio import SocketIO
from flask_mail import Mail, Message

# 主框架
async_mode = None
thread = None
app = Flask(__name__)
app.secret_key = 'lisenzzz'
socketio = SocketIO(app, async_mode=async_mode)
connection = sqlite3.connect("logindata.db", check_same_thread=False)
cur = connection.cursor()
# cur.execute("delete from udata where user = 'Jagger'")

cur.execute('CREATE TABLE IF NOT EXISTS udata (user varchar(128) PRIMARY KEY, number varchar(11), '
            'mail varchar(32), unit varchar(32), password varchar(128))')
app.config['MAIL_SERVER'] = 'smtp.qq.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = '3517717683@qq.com'
app.config['MAIL_PASSWORD'] = 'psbpzckhcmqncibf'
mail = Mail(app)

@app.route('/checkUser', methods=["POST"])
def checkUser():
    if request.method == 'POST':
        sqlite3.connect('logindata.db')
        requestArgs = request.values
        user = requestArgs.get('user')
        cur.execute("select * from udata where user = " + "'" + user + "'")
        result = cur.fetchone()
        if result is None:
            return jsonify({'isExist': False})
        elif result is not None:
            return jsonify({'isExist': True})


@app.route('/forget', methods=["GET", "POST"])
def forget():
    if request.method == 'GET':
        return render_template('forget.html')
    elif request.method == 'POST':
        requestArgs = request.values
        new = requestArgs.get('password')
        user = requestArgs.get('user')
        cur.execute("update udata set password='" + new + "' where user='" + user + "'")
        connection.commit()
        return jsonify({'isSuccess': 1})


@app.route('/register', methods=["GET", "POST"])
def register():
    if request.method == 'GET':
        return render_template('register.html')
    elif request.method == 'POST':
        requestArgs = request.values
        user = requestArgs.get('user')
        password = requestArgs.get('password')
        number = requestArgs.get('number')
        unit = requestArgs.get('unit')
        mail = requestArgs.get('mail')
        str = "'" + user + "'" + ",'" + number + "'," + "'" + mail + "'" + "," \
              + "'" + unit + "'" + "," + "'" + password + "'"
        cur.execute('insert into udata (user,number,mail,unit,password) values (' + str + ")")
        connection.commit()
        return jsonify({'isSuccess': 1})


@app.route('/send', methods=["POST"])
def send():
    requestArgs = request.values
    dirMail = requestArgs.get('mail')
    user = requestArgs.get('user')
    # print(user)

    if user is not None:
        cur.execute("select * from udata where user = " + "'" + user + "'")
        # print("select * from udata where user = " + "'" + user + "'")
        result = cur.fetchone()
        # print("mail" + result[2])
        if result != None:
            if result[2] != dirMail:
                return jsonify({'ischecked': 0})
    verificationList = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', 'a', 'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h',
                        'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 's', 't', 'x', 'y', 'z', 'A', 'B', 'C', 'D',
                        'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'S', 'T', 'X', 'Y', 'Z']
    veriCode = ''
    for i in range(4):
        veriCode += verificationList[random.randint(0, len(verificationList) - 1)]
    msg = Message("可视化平台验证码", sender="3517717683@qq.com", recipients=[dirMail])
    msg.body = veriCode
    mail.send(msg)
    return jsonify({'code': veriCode, 'ischecked': 1})

# 登录后台服务
@app.route('/login', methods=["GET", "POST"])
def login():
    if request.method == 'GET':
        sqlite3.connect('logindata.db')
        return render_template('login.html')
    elif request.method == 'POST':
        requestArgs = request.values
        user = requestArgs.get('user')
        password = requestArgs.get('password')
        # print("password" + password)
        cur.execute("select * from udata where user = " + "'" + user + "'")
        result = cur.fetchone()  # 没找到为None, 否则返回对应的元组
        cur.execute("select * from udata where password = " + "'" + password + "'")
        # p = cur.fetchone()  # 返回的是三元组，p[0]是需要的值
        # for i in range(5):
        #     print(result[i] + "result" )
        check = {'userInfo': -1, 'passwordInfo': -1}
        if result is None:
            check['userInfo'] = -1
        elif result is not None:
            check['userInfo'] = 0
            if password == result[4]:
                check['passwordInfo'] = 1
            elif password != result[4]:
                check['passwordInfo'] = 0
        check = json.dumps(check)
        return jsonify({'check': check})


@app.route('/fun', methods=["POST"])
def fun():
    requestArgs = request.values
    user = requestArgs.get('userName')
    cur.execute("select * from udata where user = " + "'" + user + "'")
    result = cur.fetchone()
    return jsonify({'user': result})

# 主界面跳转服务
@app.route('/')
def index():
    return render_template('index.html')

# 获得已登录用户信息服务
@app.route('/getUserInfo', methods=["POST"])
def getUserInfo():
    userInfo = {}
    if request.method == 'POST':
        sqlite3.connect('logindata.db')
        requestArgs = request.values
        user = requestArgs.get('user')
        # print(user)
        # if user is None:
        #     return jsonify({'isSuccess': 0})
        cur.execute("select * from udata where user = " + "'" + user + "'")
        result = cur.fetchone()
        # print(type(result[2]), type(result[3]),type(result[1]))
        if result is None:
            return jsonify({'isSuccess': 0})
        elif result[0] == user:
            userInfo['user'] = result[0]
            userInfo['phoneNumber'] = result[1]
            userInfo['mail'] = result[2]
            userInfo['unit'] = result[3]
            userInfo['isSuccess'] = 1
            newData = json.dumps(userInfo)  # json.dumps封装
            return newData

# 使用说明，内嵌子页面跳转服务
@app.route('/introduce')
def introduce():
    '''
    TODO
    :return:
    '''
    test = "使用说明界面"
    # 把需要的数据给对应的页面
    return render_template('introduce.html', test=test)

# 关于我们，内嵌子页面跳转服务
@app.route('/aboutUs')
def aboutUs():
    '''
    TODO
    :return:
    '''
    return render_template('aboutUs.html')

# 单条微博情感分析，后端跳转服务
@app.route('/Request_Extract')
def Request_Extract():
    return render_template('Request_Extract.html')

#request_extract
# 诉求抽取，后端页面跳转服务
@app.route('/LawsuitExtract', methods=["POST"])
def LawsuitExtract():
    newData = {}
    # 获取前端请求的数据
    content = request.form.get('content')
    content = content.replace('\n', '')
    content = content.replace('\r', '')
    content = content.replace('\t', '')
    content = content.replace(' ', '')
    #算法处理
    if request_extract.emergency_degree_classification(content, request_extract.emergency_word) == True:
        newData['degree_of_urgency'] = "紧急"
    else:
        newData['degree_of_urgency'] = "一般"
    newData['issue'] = request_extract.get_request(content, request_extract.request_word, request_extract.request_double_word)
    temp_request = request_extract.get_keyinfo(content)
    if newData['issue'] == '':
        newData['issue'] = request_extract.get_request_by_keyword(content)
    newData['keyword'] = temp_request['keywords']
    newData['organization'] = temp_request['org']
    newData['address'] = temp_request['location']
    #[1~5]分别对应['可忽略危险', '临界危险', '一般危险', '破坏性危险', '毁灭性危险']
    newData['degree_of_dangerous'] = request_extract.dangerous_degree_classification(content, request_extract.dangerous_word)
    newData['department'] = predict.predict(content)
    # 填补情感分析
    if content != '':
        newData['content_err'] = 1
        newData['senti_value'] = emotionalAnalysisOfSingleData.sen_value(emotionalAnalysisOfSingleData.clearTxt(content))
    else:
        newData['content_err'] = 0
        newData['senti_value'] = 0
    # print(temp_request)
    # print(content)
    return json.dumps(newData)

#写入预测结果到第7列
def write_predict_result(dir, name, save_name, request_Columns, department_Columns, result_Colunms, batch_size=64):
    file_path = os.path.join(dir, name)
    save_path = os.path.join(dir, save_name)

    # 打开文件
    workbook = openpyxl.load_workbook(file_path)
    sheetnames = workbook.get_sheet_names()
    sheet = workbook.get_sheet_by_name(sheetnames[0])

    ## 或者如果你只是想创建一张空表
    copy_workbook = openpyxl.Workbook()
    # 创建一个sheet
    copy_sheet = copy_workbook.create_sheet(index=0)

    # 写入一个值，括号内分别为行数、列数、内容
    copy_sheet.cell(1, 1).value = "登记时间"
    copy_sheet.cell(1, 2).value = "诉求"
    copy_sheet.cell(1, 3).value = "目的代码"
    copy_sheet.cell(1, 4).value = "目的"
    copy_sheet.cell(1, 5).value = "分类代码"
    copy_sheet.cell(1, 6).value = "分类"
    copy_sheet.cell(1, 7).value = "预测分类"
    copy_sheet.cell(1, 8).value = "紧急程度"
    copy_sheet.cell(1, 9).value = "危险程度"
    copy_sheet.cell(1, 10).value = "核心词汇"
    copy_sheet.cell(1, 11).value = "诉求问题"
    copy_sheet.cell(1, 12).value = "相关机构"
    copy_sheet.cell(1, 13).value = "关联地址"
    # 遍历
    rows = sheet.max_row
    clos = sheet.max_column
    print('rows:', rows)
    print('clos:', clos)
    for row in range(2, rows + 1, batch_size):
        #防止最后一批数据越界
        if row+batch_size > rows+1:
            batch_size = rows + 1 - row
        #一次性获取一批诉求
        request_list = []
        for index in range(row, row + batch_size):
            # 获取逐条投诉数据
            request = sheet.cell(index, request_Columns).value
            request = request.replace('\n', '')
            request = request.replace('\r', '')
            request = request.replace('\t', '')
            request = request.replace(' ', '')
            request_list.append(request)
        #批预测
        predict_department_list = predict.predict_list(request_list)
        #判断输入输出数量是否跟批次对应上
        assert len(request_list) == batch_size and len(predict_department_list) == batch_size
        #一次性写入一批结果
        for index in range(row, row + batch_size):
            copy_sheet.cell(index, 1).value = sheet.cell(index, 1).value
            copy_sheet.cell(index, 2).value = sheet.cell(index, 2).value
            copy_sheet.cell(index, 3).value = sheet.cell(index, 3).value
            copy_sheet.cell(index, 4).value = sheet.cell(index, 4).value
            copy_sheet.cell(index, 5).value = sheet.cell(index, 5).value
            copy_sheet.cell(index, 6).value = sheet.cell(index, 6).value
            # 7列为预测分类结果
            copy_sheet.cell(index, result_Colunms).value = predict_department_list[index - row]

            copy_sheet.cell(index, 8).value = sheet.cell(index, 8).value
            copy_sheet.cell(index, 9).value = sheet.cell(index, 9).value
            copy_sheet.cell(index, 10).value = sheet.cell(index, 10).value
            copy_sheet.cell(index, 11).value = sheet.cell(index, 11).value
            copy_sheet.cell(index, 12).value = sheet.cell(index, 12).value
            copy_sheet.cell(index, 13).value = sheet.cell(index, 13).value
        print(row)

    copy_workbook.save(save_path)
    print('finish...')

if __name__ == '__main__':
    # #跑出预测结果到表中第7列
    # dir = './static/data/'
    # name = 'excel1_with_results.xlsx'
    # save_name = 'excel2_with_results.xlsx'
    # request_Columns = 2
    # department_Columns = 6
    # result_Colunms = 7
    # batch_size = 32
    # write_predict_result(dir, name, save_name, request_Columns, department_Columns, result_Colunms, batch_size)
    #总运行
    app.run()
